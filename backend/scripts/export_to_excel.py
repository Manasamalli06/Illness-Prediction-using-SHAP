import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def export_to_excel():
    # Paths
    base_dir = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(base_dir, '../data/augmented_medical_data.csv')
    xlsx_path = os.path.join(base_dir, '../data/augmented_medical_data.xlsx')

    print("Reading CSV file...")
    df = pd.read_csv(csv_path)
    print(f"Loaded {len(df)} rows and {len(df.columns)} columns.")

    # Create workbook manually for clean output
    wb = Workbook()
    ws = wb.active
    ws.title = "Medical Dataset"

    # --- Header Style ---
    header_fill   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font   = Font(color="FFFFFF", bold=True, size=11)
    header_align  = Alignment(horizontal="center", vertical="center")
    thin_border   = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Row Styles ---
    low_risk_fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green
    high_risk_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # light red
    alt_fill       = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # light grey

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        risk_label = row[-1]  # Last column is Risk_Label
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            # Color-code rows by risk
            if risk_label == 'High Risk':
                cell.fill = high_risk_fill
            elif risk_label == 'Low Risk':
                cell.fill = low_risk_fill
            else:
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

    # --- Auto-fit column widths ---
    col_widths = {
        'Age': 8, 'Gender': 14, 'BMI': 10,
        'Systolic_BP': 14, 'Glucose': 12,
        'Body_Temp': 12, 'Risk_Label': 14
    }
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

    # Freeze top header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Save
    wb.save(xlsx_path)
    print(f"\n✅ Excel file saved successfully!")
    print(f"   Path: {xlsx_path}")
    print(f"   Rows: {len(df)}")
    print(f"   Columns: {', '.join(df.columns.tolist())}")

if __name__ == "__main__":
    export_to_excel()
